from openpyxl import load_workbook
import os
import logging


class GetExcel:
    def __init__(self):
        self.res_table = list()
        self.keys = ['name', 'surname', 'patronymic', 'mail']
        self.files = os.listdir('tables/')
        print(self.files)
        logging.basicConfig(filename="table.log",
                            level=logging.INFO,
                            datefmt='%I:%M:%S',
                            format='%(asctime)s %(message)s')

    def get_data(self, excel_file):

        logging.info('File %s' % excel_file)
        wb = load_workbook('tables/%s' % excel_file)
        ws = wb.active
        logging.info('Rows in file: %d' % len(list(ws)))

        for row in ws.iter_rows():
            cell = [cell.value for cell in row]
            values = cell[0].split()  # TODO: names cell
            values.append(cell[1])  # TODO: mail cell
            res_cell = {key: name for key, name in zip(self.keys, values)}
            self.res_table.append(res_cell)

        logging.info('Taken rows: %d' % (len(self.res_table)))


def get_all_from_files():
    excel = GetExcel()
    for file in excel.files:
        if file[0] != '.':
            excel.get_data(file)
    return excel.res_table

def get_part_from_tables(head_file=0, tail_file=1, rows=1):
    excel = GetExcel()
    if tail_file > len(excel.files):
        return 'error'
    for i in range(head_file, tail_file):
        if excel.files[i][0] != '.':
            excel.get_data(excel.files[i])

    return excel.res_table

def get_one_file(filename):
    excel = GetExcel()
    if filename in excel.files:
        excel.get_data(filename)
    return excel.res_table
